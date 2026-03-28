from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

OUTPUT_PATH = "/workspace/outputs/Comms_Measurement_Framework.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY       = "1B3A5C"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MID_GRAY   = "D9D9D9"
ACCENT_BLUE= "2E75B6"
RED_FILL   = "FF0000"
AMBER_FILL = "FFC000"
GREEN_FILL = "70AD47"
DARK_RED   = "C00000"
DARK_AMBER = "9C5700"
DARK_GREEN = "375623"

# ── Reusable style helpers ────────────────────────────────────────────────────
def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def white_fill():
    return PatternFill("solid", fgColor=WHITE)

def mid_gray_fill():
    return PatternFill("solid", fgColor=MID_GRAY)

def thin_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row, cols, fill=None):
    f = fill or navy_fill()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = hdr_font()
        cell.alignment = center()
        cell.border = thin_border()

def apply_data_row(ws, row, cols, alt=False):
    f = gray_fill() if alt else white_fill()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = body_font()
        cell.alignment = left()
        cell.border = thin_border()

def write_row(ws, row, values, cols=None):
    for c, v in enumerate(values, 1):
        ws.cell(row=row, column=c, value=v)
    if cols and len(values) < cols:
        for c in range(len(values) + 1, cols + 1):
            ws.cell(row=row, column=c, value="")

def freeze_and_set_zoom(ws, cell="A2"):
    ws.freeze_panes = cell

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Dashboard Summary
# ══════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard Summary"
    ws.sheet_view.showGridLines = False

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "Communications Plan Measurement & Effectiveness Framework"
    ws["A1"].font = Font(name="Arial", bold=True, size=18, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    ws["A2"] = "Internal Comms Plan — Leadership Strategy Rollout"
    ws["A2"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=ACCENT_BLUE)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:E3")
    ws["A3"] = "Reporting Period: March 2026"
    ws["A3"].font = body_font(size=11, bold=True, color=NAVY)
    ws["A3"].fill = gray_fill()
    ws["A3"].alignment = center()
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 10  # spacer

    # ── Section label ────────────────────────────────────────────────────────
    ws.merge_cells("A5:E5")
    ws["A5"] = "AUDIENCE HEALTH SUMMARY"
    ws["A5"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A5"].fill = navy_fill()
    ws["A5"].alignment = center()
    ws.row_dimensions[5].height = 22

    # ── Table headers ────────────────────────────────────────────────────────
    headers = [
        "Audience Group",
        "Overall Health Score",
        "Key Risk",
        "Priority Action",
    ]
    # 5 cols but we only use 4 data cols; keep col E as buffer
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="2E4F7F")
        cell.font = hdr_font(size=10)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[6].height = 22

    rows = [
        (
            "Corporate Employees",
            "To be measured",
            "Low engagement with strategic narrative",
            "Launch targeted intranet content",
        ),
        (
            "Frontline / Operations Employees",
            "To be measured",
            "Disconnect between corporate messaging and daily experience",
            "Manager-led sessions with local context",
        ),
        (
            "People Managers",
            "To be measured",
            "Burnout and insufficient preparation",
            "Immediate coaching investment",
        ),
    ]

    for i, row_data in enumerate(rows):
        r = 7 + i
        alt = i % 2 == 1
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = gray_fill() if alt else white_fill()
            cell.font = body_font(size=10)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[r].height = 36

    # Blank col E for all data rows
    for r in range(6, 10):
        cell = ws.cell(row=r, column=5, value="")
        cell.fill = gray_fill() if (r % 2 == 1) else white_fill()
        cell.border = thin_border()

    # ── Footer note ──────────────────────────────────────────────────────────
    ws.row_dimensions[10].height = 10
    ws.merge_cells("A11:E11")
    ws["A11"] = (
        "Note: Health scores will be populated after first 30-day measurement cycle. "
        "See 'Success Metrics' sheet for full KPI tracking."
    )
    ws["A11"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[11].height = 28

    # ── Column widths ────────────────────────────────────────────────────────
    set_col_widths(ws, {
        "A": 32, "B": 22, "C": 46, "D": 44, "E": 6,
    })

    freeze_and_set_zoom(ws, "A2")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Success Metrics
# ══════════════════════════════════════════════════════════════════════════════
def build_metrics(wb):
    ws = wb.create_sheet("Success Metrics")
    ws.sheet_view.showGridLines = False

    # ── Title ────────────────────────────────────────────────────────────────
    COLS = 10
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    ws["A1"] = "Success Metrics — Communications Effectiveness Measurement"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 34

    # ── Headers ──────────────────────────────────────────────────────────────
    headers = [
        "Metric ID", "Metric Name", "Description",
        "Audience Group", "Data Source",
        "Baseline (Current)", "Target (30 Days)", "Target (90 Days)",
        "Measurement Frequency", "Owner",
    ]
    write_row(ws, 2, headers)
    apply_header_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 36

    # ── Data ─────────────────────────────────────────────────────────────────
    data = [
        ("M-001", "Email Open Rate — Corporate",
         "Percentage of corporate employees who open leadership email communications",
         "Corporate", "Email platform analytics (e.g., Outlook Insights)",
         "42%", "60%", "70%", "Per send", "Comms Lead"),
        ("M-002", "Email Read Rate — Corporate",
         "Percentage of corporate employees who read >50% of email content",
         "Corporate", "Email platform analytics",
         "28%", "45%", "60%", "Per send", "Comms Lead"),
        ("M-003", "Email Open Rate — Frontline",
         "Percentage of frontline employees who open communications via email or app",
         "Frontline", "Email / workforce app analytics",
         "25%", "45%", "60%", "Per send", "Ops Comms"),
        ("M-004", "Email Read Rate — Frontline",
         "Percentage of frontline employees who fully read communications",
         "Frontline", "Email / workforce app analytics",
         "15%", "35%", "50%", "Per send", "Ops Comms"),
        ("M-005", "Intranet Hub Page Views",
         "Total page views on the dedicated leadership strategy intranet hub",
         "All", "Intranet analytics (SharePoint / Confluence)",
         "0 (new)", "2,500", "5,000", "Weekly", "Digital Comms"),
        ("M-006", "Intranet Hub Unique Visitors",
         "Unique employee visitors to the leadership strategy intranet hub",
         "All", "Intranet analytics",
         "0 (new)", "800", "1,500", "Weekly", "Digital Comms"),
        ("M-007", "FAQ Page Engagement Rate",
         "Percentage of hub visitors who click into the FAQ section",
         "All", "Intranet analytics",
         "N/A", "30%", "50%", "Weekly", "Comms Lead"),
        ("M-008", "FAQ Page Average Time on Page",
         "Average time employees spend reading the FAQ page (signals depth of engagement)",
         "All", "Intranet analytics",
         "N/A", "2 min", "3 min", "Weekly", "Digital Comms"),
        ("M-009", "Manager Meeting Completion Rate",
         "Percentage of team meetings completed by managers within the rollout window",
         "Managers", "Manager reporting / HR system",
         "N/A", "80% by Day 7", "95% by Day 30", "Weekly", "HR Lead"),
        ("M-010", "Employee Sentiment — Understanding",
         "Pulse survey score: employees report understanding the strategic direction (1–10 scale)",
         "All", "Pulse survey platform",
         "Baseline to be set at launch", "6.5 / 10", "7.5 / 10", "Bi-weekly", "HR Lead"),
        ("M-011", "Employee Sentiment — Confidence",
         "Pulse survey score: employees feel confident about their role in the strategy",
         "All", "Pulse survey platform",
         "Baseline to be set at launch", "6.0 / 10", "7.0 / 10", "Bi-weekly", "HR Lead"),
        ("M-012", "Employee Sentiment — Trust in Leadership",
         "Pulse survey score: employees trust leadership communication (1–10 scale)",
         "All", "Pulse survey platform",
         "Baseline to be set at launch", "6.0 / 10", "7.5 / 10", "Monthly", "CHRO"),
        ("M-013", "Manager Confidence Score",
         "Managers self-report confidence in delivering key strategic messages (1–10 scale)",
         "Managers", "Manager feedback form / pulse",
         "Baseline to be set at kickoff", "7.0 / 10", "8.5 / 10", "Bi-weekly", "HR Lead"),
        ("M-014", "Voluntary Attrition Rate Change",
         "Month-over-month change in voluntary attrition versus pre-rollout baseline",
         "All", "HRIS (Workday / SAP SuccessFactors)",
         "Current monthly attrition rate", "+/- 0% vs baseline", "No increase vs baseline", "Monthly", "CHRO"),
        ("M-015", "Questions Submitted via Feedback Channels",
         "Volume and themes of questions submitted through 'Ask Leadership' or feedback tools",
         "All", "Intranet / email submission tool",
         "0 (new channel)", "50 questions", "100 questions", "Weekly", "Comms Lead"),
        ("M-016", "Manager Coaching Program Enrollment",
         "Percentage of people managers enrolled in the leadership messaging coaching program",
         "Managers", "LMS / HR training records",
         "0%", "70%", "100%", "Weekly", "L&D Lead"),
        ("M-017", "eNPS Shift (Employee Net Promoter Score)",
         "Change in eNPS score versus pre-rollout benchmark — measures employee advocacy",
         "All", "eNPS pulse survey",
         "Pre-rollout score to be captured", "+5 points", "+10 points", "Quarterly", "CHRO"),
        ("M-018", "Internal Search Trends",
         "Top search terms on intranet — identifies gaps, fears, and information needs",
         "All", "Intranet search analytics",
         "Baseline snapshot at launch", "Strategy-related terms in top 10", "FAQ terms declining (needs met)", "Weekly", "Digital Comms"),
        ("M-019", "Engagement Survey Scores — Next Cycle",
         "Scores on 'strategic direction', 'leadership communication', and 'change management' in next annual engagement survey",
         "All", "Annual engagement survey",
         "Previous survey scores", "+5% on target questions", "+10% on target questions", "Annual (next cycle)", "CHRO"),
    ]

    for i, row_data in enumerate(data):
        r = 3 + i
        alt = i % 2 == 1
        write_row(ws, r, list(row_data), COLS)
        apply_data_row(ws, r, COLS, alt)
        ws.row_dimensions[r].height = 44

    # Center Metric ID and Frequency cols
    for r in range(3, 3 + len(data)):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    set_col_widths(ws, {
        "A": 12, "B": 32, "C": 50, "D": 14, "E": 30,
        "F": 22, "G": 18, "H": 18, "I": 22, "J": 18,
    })
    freeze_and_set_zoom(ws, "A3")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Leading Indicators
# ══════════════════════════════════════════════════════════════════════════════
def build_leading_indicators(wb):
    ws = wb.create_sheet("Leading Indicators")
    ws.sheet_view.showGridLines = False

    COLS = 8
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    ws["A1"] = "Leading Indicators — Early Warning System"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 34

    headers = [
        "Indicator ID", "Indicator", "What It Signals",
        "Audience", "Data Source", "Alert Threshold",
        "Response Action", "Owner",
    ]
    write_row(ws, 2, headers)
    apply_header_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 36

    data = [
        ("LI-001",
         "Spike in 'layoffs', 'job security', or 'redundancy' internal search terms",
         "Employee fear and anxiety about job security; narrative not landing as intended",
         "All",
         "Intranet search analytics",
         "20%+ week-over-week increase in these search terms",
         "Immediately update FAQ with direct job security messaging; deploy manager talking points; consider executive video message",
         "Comms Lead"),
        ("LI-002",
         "Manager meeting completion rate below 60% at end of Week 1",
         "Manager resistance, scheduling barriers, or lack of enablement — risk of message vacuum",
         "Managers",
         "Manager reporting / HR system",
         "< 60% completion by end of Day 7",
         "Escalate to VP HR and department heads; deploy reminder and support pack; offer 1:1 coaching to struggling managers",
         "HR Lead"),
        ("LI-003",
         "Executive email open rate below 50%",
         "Disengagement from leadership communications; employees may be tuning out",
         "Corporate",
         "Email platform analytics",
         "< 50% open rate on any send",
         "Deploy follow-up via alternative channel (Teams, Intranet banner, manager cascade); review subject line and send timing",
         "Comms Lead"),
        ("LI-004",
         "Spike in HR inquiries related to restructuring, role changes, or org design",
         "Narrative failure — employees not getting answers through official channels; fear escalating",
         "All",
         "HR inquiry tracking (HR case management system)",
         "25%+ increase in restructuring-related inquiries week-over-week",
         "Deploy targeted clarification messaging; update FAQ; brief HR partners with holding statements; escalate to leadership if sustained",
         "HR Lead"),
        ("LI-005",
         "Decline in intranet hub return visits (repeat visitors dropping)",
         "Content is not resonating or updating frequently enough to drive ongoing engagement",
         "All",
         "Intranet analytics",
         "Return visitor rate drops below 30% in Week 2+",
         "Refresh hub with team-specific stories, manager spotlights, and Q&A responses; add new content weekly",
         "Digital Comms"),
        ("LI-006",
         "Manager feedback forms indicate inability to answer team questions",
         "Training gap — managers not sufficiently equipped to handle Q&A",
         "Managers",
         "Manager feedback form (post-session)",
         "30%+ of managers report difficulty answering questions",
         "Deploy supplemental manager briefing document within 24 hours; schedule group Q&A call with senior leader; update FAQ",
         "HR Lead / L&D Lead"),
        ("LI-007",
         "Social media / Blind platform sentiment tracking negative themes",
         "External reputation risk; internal concerns leaking publicly; talent brand damage",
         "All (external signal)",
         "Blind, LinkedIn, Glassdoor, social listening tools",
         "Any sustained (3+ days) negative theme tied to the announcement",
         "Activate external comms protocol; review internal messaging gaps; consider proactive employer brand response",
         "Comms Lead / PR"),
        ("LI-008",
         "Attrition intent survey spike (10%+ above baseline)",
         "Retention crisis emerging — employees actively considering leaving as a result of comms",
         "All",
         "Pulse survey / attrition intent question",
         "Attrition intent 10%+ above pre-announcement baseline",
         "Escalate immediately to CHRO; activate retention risk protocol; deploy personalised manager conversations; review comp and role clarity messaging",
         "CHRO"),
        ("LI-009",
         "Frontline shift briefings not completed or poorly attended",
         "Frontline employees receiving no comms — risk of rumour mill and misinformation",
         "Frontline",
         "Ops manager reporting / attendance records",
         "< 50% of shifts receive briefing in Week 1",
         "Escalate to Operations Director; deploy printed briefing cards and digital posters; adjust briefing format to fit shift patterns",
         "Ops Comms"),
        ("LI-010",
         "Low FAQ page engagement despite high hub traffic",
         "Employees visiting hub but not finding answers — content architecture or relevance issue",
         "All",
         "Intranet analytics (hub vs FAQ click-through)",
         "FAQ click-through below 20% of hub visitors",
         "Redesign hub navigation to surface FAQ more prominently; add 'top questions this week' widget; review FAQ content relevance",
         "Digital Comms"),
        ("LI-011",
         "Decline in manager confidence scores between Week 1 and Week 2 pulse",
         "Manager support measures not working; increased pressure or confusion post-launch",
         "Managers",
         "Manager pulse survey",
         "Confidence score drops more than 1 point week-over-week",
         "Schedule urgent peer learning session; provide additional talking points; ensure senior leader visibility with manager cohort",
         "HR Lead"),
        ("LI-012",
         "High volume of 'Ask Leadership' questions on same topic",
         "Identifies a specific gap in messaging — employees not satisfied with existing content on that topic",
         "All",
         "Ask Leadership submission tool",
         "5+ questions on same topic within 48 hours",
         "Prioritise published response on intranet within 24 hours; update FAQ; consider brief video response from relevant leader",
         "Comms Lead"),
    ]

    for i, row_data in enumerate(data):
        r = 3 + i
        alt = i % 2 == 1
        write_row(ws, r, list(row_data), COLS)
        apply_data_row(ws, r, COLS, alt)
        ws.row_dimensions[r].height = 60

    for r in range(3, 3 + len(data)):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    set_col_widths(ws, {
        "A": 12, "B": 42, "C": 40, "D": 12,
        "E": 32, "F": 34, "G": 50, "H": 18,
    })
    freeze_and_set_zoom(ws, "A3")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Feedback Loops
# ══════════════════════════════════════════════════════════════════════════════
def build_feedback_loops(wb):
    ws = wb.create_sheet("Feedback Loops")
    ws.sheet_view.showGridLines = False

    COLS = 8
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    ws["A1"] = "Feedback Loops — Capturing and Actioning Employee Input"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 34

    headers = [
        "Loop ID", "Feedback Mechanism", "Frequency", "Audience",
        "Input Captured", "How Input Is Used", "Escalation Path", "Owner",
    ]
    write_row(ws, 2, headers)
    apply_header_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 36

    data = [
        ("FL-001",
         "Post-Meeting Pulse Survey",
         "After each manager-led team session (Week 1 and Week 2)",
         "All Employees",
         "Clarity of message, Q&A satisfaction, sentiment score, open-text concerns",
         "Themes fed into next manager briefing pack; unresolved questions added to FAQ; Comms team updates messaging where gaps identified",
         "If >30% negative sentiment: escalate to HR Lead and Comms Lead within 24 hours",
         "HR Lead"),
        ("FL-002",
         "Intranet Comment and Reaction Tracking",
         "Real-time / reviewed weekly",
         "All Employees",
         "Reactions (likes, helpful, confused), comments, shares, content save rates",
         "Content team reviews weekly; negative reactions or confusion comments trigger content update within 48 hours",
         "Sustained negative reaction pattern: escalate to Comms Lead; offensive comments: escalate to HR",
         "Digital Comms"),
        ("FL-003",
         "Manager Feedback Form (Post-Discussion)",
         "Weekly during rollout (Weeks 1–4)",
         "People Managers",
         "Questions they couldn't answer, team mood assessment, support needs, timing issues",
         "Aggregated and sent to Comms Lead and HR Lead weekly; informs next manager briefing and FAQ updates",
         "If >25% managers flagging same concern: convene urgent manager Q&A call within 48 hours",
         "HR Lead"),
        ("FL-004",
         "Monthly Employee Sentiment Pulse",
         "Monthly (months 1–3 post-launch)",
         "All Employees",
         "Understanding, confidence, trust, alignment scores (quantitative); open-text concerns (qualitative)",
         "Results shared with senior leadership within 5 business days; comms strategy adjusted based on trends; fed into Leading Indicators tracker",
         "Any dimension dropping below threshold: immediate Comms Lead and CHRO review; response plan within 1 week",
         "CHRO / HR Lead"),
        ("FL-005",
         "Quarterly Engagement Survey Integration",
         "Quarterly (next scheduled cycle)",
         "All Employees",
         "Formal engagement scores on: strategic direction, leadership communication, change management, manager effectiveness",
         "Compared against pre-rollout benchmark; formal report to leadership team; long-term strategy adjustment",
         "Sustained decline in engagement scores: escalate to EXCO; commission independent employee listening sessions",
         "CHRO"),
        ("FL-006",
         "'Ask Leadership' Question Submission Tracking",
         "Ongoing / reviewed weekly",
         "All Employees",
         "Employee questions submitted via intranet or email; volume, themes, and sentiment",
         "Top 5 questions answered publicly on intranet within 5 business days; patterns inform FAQ updates and future comms content",
         "High volume on single topic (5+ in 48 hours): immediate FAQ update and leadership response required",
         "Comms Lead"),
        ("FL-007",
         "HR Inquiry Categorisation and Trending",
         "Weekly (throughout rollout period)",
         "All Employees",
         "Volume and category of HR queries (restructuring, role clarity, AI impact, benefits, attrition intent)",
         "Weekly HR dashboard shared with CHRO; spikes trigger comms response; holding statements updated based on top inquiry themes",
         "25%+ spike in restructuring queries: trigger Comms Lead / CHRO review; activate narrative-clarification response plan",
         "HR Lead"),
        ("FL-008",
         "Executive Listening Session Summaries",
         "Bi-weekly (first 6 weeks), then monthly",
         "Representative Employee Groups",
         "Qualitative themes from small-group sessions with senior leaders; unfiltered employee voice",
         "Verbatim themes shared with EXCO (anonymised); directly informs leadership communication tone and content",
         "Themes suggesting fundamental trust breakdown or morale crisis: immediate EXCO escalation and response session required",
         "CHRO / Comms Lead"),
    ]

    for i, row_data in enumerate(data):
        r = 3 + i
        alt = i % 2 == 1
        write_row(ws, r, list(row_data), COLS)
        apply_data_row(ws, r, COLS, alt)
        ws.row_dimensions[r].height = 72

    for r in range(3, 3 + len(data)):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    set_col_widths(ws, {
        "A": 10, "B": 34, "C": 22, "D": 22,
        "E": 40, "F": 44, "G": 44, "H": 16,
    })
    freeze_and_set_zoom(ws, "A3")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Rollout Tracker
# ══════════════════════════════════════════════════════════════════════════════
def build_rollout_tracker(wb):
    ws = wb.create_sheet("Rollout Tracker")
    ws.sheet_view.showGridLines = False

    COLS = 8
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    ws["A1"] = "Rollout Tracker — 2-Week Communications Activity Plan"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 34

    headers = [
        "Week", "Day", "Activity", "Channel",
        "Audience", "Status", "Owner", "Notes",
    ]
    write_row(ws, 2, headers)
    apply_header_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 30

    data = [
        # Week 1
        ("Week 1", "Mon Day 1",
         "Executive all-company email — strategic direction announcement",
         "Email (Outlook)", "All Employees", "Planned", "Comms Lead",
         "Subject line pre-tested with employee panel; personalised by CEO"),
        ("Week 1", "Mon Day 1",
         "Intranet leadership strategy hub — go-live",
         "Intranet (SharePoint)", "All Employees", "Planned", "Digital Comms",
         "Hub includes: overview, FAQ, video, manager toolkit, Q&A submission"),
        ("Week 1", "Mon Day 1",
         "Manager briefing pack distributed (email + Teams)",
         "Email / Microsoft Teams", "People Managers", "Planned", "HR Lead",
         "Pack includes: talking points, FAQ, briefing guide, pulse survey link"),
        ("Week 1", "Tue Day 2",
         "Manager team briefing sessions begin (all teams)",
         "In-person / virtual team meetings", "All Employees", "Planned", "People Managers",
         "Managers use briefing pack; completion tracked via reporting form"),
        ("Week 1", "Tue Day 2",
         "Frontline shift briefings begin (operations)",
         "In-person shift briefing / printed cards", "Frontline", "Planned", "Ops Comms",
         "Shift supervisor-led; printed one-pager provided for all shifts"),
        ("Week 1", "Wed Day 3",
         "Post-meeting pulse survey deployed (employees)",
         "Pulse survey platform", "All Employees", "Planned", "HR Lead",
         "5-question survey sent to all employees who attended a team briefing"),
        ("Week 1", "Wed Day 3",
         "Manager coaching programme — cohort 1 session",
         "Virtual (Teams/Zoom)", "People Managers", "Planned", "L&D Lead",
         "1-hour live session: handling tough Q&A, reading team sentiment"),
        ("Week 1", "Thu Day 4",
         "'Ask Leadership' Q&A channel — first response batch published",
         "Intranet hub", "All Employees", "Planned", "Comms Lead",
         "Top 10 questions answered; published within 72 hours of launch"),
        ("Week 1", "Thu Day 4",
         "HR inquiry categorisation — first weekly review",
         "HR case management system", "HR Team", "Planned", "HR Lead",
         "Identify top themes; brief CHRO; update holding statements if needed"),
        ("Week 1", "Fri Day 5",
         "Manager feedback form — first submission deadline",
         "Online form (intranet)", "People Managers", "Planned", "HR Lead",
         "Managers submit: questions raised, mood assessment, support gaps"),
        ("Week 1", "Fri Day 5",
         "Week 1 leading indicators review — Comms & HR team",
         "Internal meeting (Teams)", "Comms & HR Teams", "Planned", "Comms Lead",
         "Review: open rates, hub traffic, FAQ engagement, search trends, pulse survey results"),
        ("Week 1", "Fri Day 5",
         "Intranet hub content update — refresh with week 1 Q&A themes",
         "Intranet (SharePoint)", "All Employees", "Planned", "Digital Comms",
         "Add top Q&As from week; update FAQ with any unanticipated questions"),
        # Week 2
        ("Week 2", "Mon Day 8",
         "Follow-up email to non-openers — corporate segment",
         "Email (Outlook)", "Corporate (non-openers)", "Planned", "Comms Lead",
         "Shortened version; different subject line; include direct hub link"),
        ("Week 2", "Mon Day 8",
         "Manager coaching programme — cohort 2 session",
         "Virtual (Teams/Zoom)", "People Managers", "Planned", "L&D Lead",
         "Focus on managers who flagged difficulty in Week 1 feedback form"),
        ("Week 2", "Tue Day 9",
         "Executive listening session — Cohort 1 (Corporate)",
         "Virtual small-group session", "Corporate Employees (sample)", "Planned", "CHRO",
         "15–20 employees; CEO or senior leader hosts; themes captured and shared"),
        ("Week 2", "Tue Day 9",
         "Frontline shift briefings — completion check and catch-up",
         "In-person / printed materials", "Frontline", "Planned", "Ops Comms",
         "Any missed shifts from Week 1 completed; updated Q&A cards distributed"),
        ("Week 2", "Wed Day 10",
         "Monthly sentiment pulse survey launch",
         "Pulse survey platform", "All Employees", "Planned", "HR Lead",
         "Includes: understanding, confidence, trust, and eNPS question"),
        ("Week 2", "Wed Day 10",
         "Social media / Blind sentiment review",
         "Social listening tool / manual review", "Comms Lead", "Planned", "Comms Lead",
         "Check for negative themes related to announcement; flag to PR if needed"),
        ("Week 2", "Thu Day 11",
         "FAQ page update — second batch based on 'Ask Leadership' submissions",
         "Intranet (SharePoint)", "All Employees", "Planned", "Digital Comms",
         "Prioritise any topics with 3+ submissions; tag manager-specific topics"),
        ("Week 2", "Thu Day 11",
         "HR inquiry trending report — shared with CHRO and Comms Lead",
         "Internal report (email)", "CHRO / Comms Lead", "Planned", "HR Lead",
         "2-week trend vs launch baseline; flag any emerging risk themes"),
        ("Week 2", "Fri Day 12",
         "Manager feedback form — Week 2 submission deadline",
         "Online form (intranet)", "People Managers", "Planned", "HR Lead",
         "Assess improvement vs Week 1; identify persistent gaps"),
        ("Week 2", "Fri Day 12",
         "2-week rollout retrospective — Comms, HR, and Ops teams",
         "In-person / virtual meeting", "Comms & HR & Ops Teams", "Planned", "Comms Lead",
         "Review all leading indicators, pulse results, and feedback; agree 30-day plan adjustments"),
        ("Week 2", "Fri Day 12",
         "30-day plan confirmed and communicated to stakeholders",
         "Email / Teams", "Senior Leadership", "Planned", "Comms Lead",
         "Summary of 2-week results, metrics status, and adjusted 30-day approach"),
    ]

    for i, row_data in enumerate(data):
        r = 3 + i
        alt = i % 2 == 1
        # Colour Week 1 vs Week 2 section headers
        write_row(ws, r, list(row_data), COLS)
        apply_data_row(ws, r, COLS, alt)
        ws.row_dimensions[r].height = 40

        # Bold "Week 1" / "Week 2" in column A
        cell_a = ws.cell(row=r, column=1)
        cell_a.font = Font(name="Arial", bold=True, size=10,
                           color=NAVY if row_data[0] == "Week 1" else "C00000")
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Status column colouring
        status_cell = ws.cell(row=r, column=6)
        status_cell.font = Font(name="Arial", size=10, bold=True, color="375623")
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_col_widths(ws, {
        "A": 9, "B": 14, "C": 50, "D": 28,
        "E": 22, "F": 12, "G": 16, "H": 44,
    })
    freeze_and_set_zoom(ws, "A3")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – Risk Register
# ══════════════════════════════════════════════════════════════════════════════
def build_risk_register(wb):
    ws = wb.create_sheet("Risk Register")
    ws.sheet_view.showGridLines = False

    COLS = 8
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    ws["A1"] = "Risk Register — Communications Plan Risks and Mitigations"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 34

    headers = [
        "Risk ID", "Risk Description",
        "Likelihood (H/M/L)", "Impact (H/M/L)", "Risk Score",
        "Mitigation Strategy", "Owner", "Status",
    ]
    write_row(ws, 2, headers)
    apply_header_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 36

    data = [
        ("R-001",
         "Employees interpret 'startup culture' messaging as a signal of impending layoffs or role changes",
         "H", "H", "High",
         "Pre-emptive FAQ with explicit job security messaging; manager talking points addressing this directly; review all copy before send",
         "Comms Lead", "Active"),
        ("R-002",
         "Manager burnout prevents effective and consistent message delivery to teams",
         "H", "H", "High",
         "Manager support package with ready-to-use materials; reduce meeting admin burden; coaching programme; peer support groups",
         "HR Lead", "Active"),
        ("R-003",
         "Frontline employees do not receive or engage with corporate communications due to shift patterns and screen access",
         "M", "H", "High",
         "Multi-channel distribution (printed, digital screens, shift briefings); shift-friendly formats; ops manager accountability",
         "Ops Comms", "Active"),
        ("R-004",
         "AI and automation messaging triggers widespread job security anxiety across all employee groups",
         "H", "M", "Medium",
         "Specific AI FAQ section with career development commitments; avoid jargon; include concrete role protection language",
         "Comms Lead", "Active"),
        ("R-005",
         "Executive email is perceived as tone-deaf, out-of-touch, or insufficiently empathetic",
         "M", "H", "High",
         "Pre-test email copy with diverse employee panels; include personal and empathetic framing; avoid corporate-speak",
         "Comms Lead", "Active"),
        ("R-006",
         "Feedback loops fail to surface genuine employee concerns due to low trust or fear of visibility",
         "M", "M", "Medium",
         "Ensure anonymous feedback channels; train managers to encourage honest input; communicate how feedback is used",
         "HR Lead", "Active"),
        ("R-007",
         "Inconsistent message delivery across managers creates confusion and contradictory narratives",
         "M", "H", "High",
         "Single-source briefing pack with locked key messages; cascade check-in call before sessions begin; manager Q&A hotline",
         "HR Lead / Comms Lead", "Active"),
        ("R-008",
         "Communications plan does not reach or land with geographically dispersed or remote employees",
         "M", "M", "Medium",
         "Ensure digital-first channels; regional manager cascade; intranet hub accessible globally; translated materials if needed",
         "Digital Comms", "Active"),
        ("R-009",
         "Negative sentiment leaks externally via Blind, social media, or press — reputational damage",
         "L", "H", "Medium",
         "Social listening activated day 1; PR on standby for reactive response; ensure internal comms is sufficiently reassuring",
         "Comms Lead / PR", "Active"),
        ("R-010",
         "Leadership team is not aligned on key messages — contradictions emerge in skip-level or direct interactions",
         "L", "H", "Medium",
         "Pre-launch leadership alignment session; single-source messaging document; Comms Lead approves all executive communications",
         "Comms Lead", "Active"),
    ]

    for i, row_data in enumerate(data):
        r = 3 + i
        alt = i % 2 == 1
        write_row(ws, r, list(row_data), COLS)
        apply_data_row(ws, r, COLS, alt)
        ws.row_dimensions[r].height = 52

        cell_a = ws.cell(row=r, column=1)
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        cell_a.font = Font(name="Arial", bold=True, size=10)

        likelihood_cell = ws.cell(row=r, column=3)
        impact_cell     = ws.cell(row=r, column=4)
        score_cell      = ws.cell(row=r, column=5)

        for cell in [likelihood_cell, impact_cell, score_cell]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", bold=True, size=10)

        score = row_data[4]
        if score == "High":
            score_cell.fill = PatternFill("solid", fgColor="FF0000")
            score_cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        elif score == "Medium":
            score_cell.fill = PatternFill("solid", fgColor="FFC000")
            score_cell.font = Font(name="Arial", bold=True, size=10, color="000000")
        else:
            score_cell.fill = PatternFill("solid", fgColor="70AD47")
            score_cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)

        status_cell = ws.cell(row=r, column=8)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.font = Font(name="Arial", bold=True, size=10, color="375623")

    # Legend
    legend_row = 3 + len(data) + 1
    ws.merge_cells(f"A{legend_row}:H{legend_row}")
    ws[f"A{legend_row}"] = (
        "Risk Score Legend:   HIGH = High Likelihood × High or Medium Impact   |   "
        "MEDIUM = Medium Likelihood × Medium/High Impact or High Likelihood × Low Impact   |   "
        "LOW = Low Likelihood × Low Impact"
    )
    ws[f"A{legend_row}"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws[f"A{legend_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f"A{legend_row}"].fill = gray_fill()
    ws.row_dimensions[legend_row].height = 28

    set_col_widths(ws, {
        "A": 10, "B": 52, "C": 16, "D": 14,
        "E": 12, "F": 50, "G": 20, "H": 12,
    })
    freeze_and_set_zoom(ws, "A3")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    build_dashboard(wb)
    build_metrics(wb)
    build_leading_indicators(wb)
    build_feedback_loops(wb)
    build_rollout_tracker(wb)
    build_risk_register(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
